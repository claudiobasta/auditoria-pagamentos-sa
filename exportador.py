"""
Geração do xlsx final no formato exato esperado pelo financeiro.
Replica a formatação observada em "Pagamentos - Equipe externa.xlsx".
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


HEADER_FILL = PatternFill('solid', start_color='FFEAF2DF')
HEADER_FONT = Font(name='Segoe UI', size=8, bold=True, color='FF000000')
DATA_FONT = Font(name='Calibri', size=11)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_JUSTIFY = Alignment(horizontal='justify', vertical='center')

BORDER_MEDIUM = Border(
    top=Side(style='medium'),
    bottom=Side(style='medium'),
    left=Side(style='medium'),
    right=Side(style='medium'),
)

COL_WIDTHS = {
    'A': 10.66, 'B': 47.55, 'C': 23.0, 'D': 12.10, 'E': 14.0,
    'F': 16.55, 'G': 60.0, 'H': 23.88, 'I': 16.44,
}

HEADER_CENTER_COLS = {'E', 'F', 'H'}  # data registro, vencimento, chave pix
FORMATO_MOEDA = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@_-'
FORMATO_DATA = 'mm-dd-yy'


def gerar_xlsx_financeiro(df_pagamentos):
    """Recebe DataFrame final e devolve bytes do xlsx formatado."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Equipe externa'

    # Cabeçalho
    headers = list(df_pagamentos.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER_MEDIUM
        col_letter = get_column_letter(col_idx)
        cell.alignment = ALIGN_CENTER if col_letter in HEADER_CENTER_COLS else ALIGN_JUSTIFY

    # Dados
    for row_idx, (_, row) in enumerate(df_pagamentos.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            col_letter = get_column_letter(col_idx)
            if col_letter == 'D':
                cell.number_format = FORMATO_MOEDA
            elif col_letter in ('E', 'F'):
                cell.number_format = FORMATO_DATA

    # Larguras
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
